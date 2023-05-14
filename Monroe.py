import openpyxl
import os
import sys
import tkinter
from tkinter import messagebox

#carga la planilla
try:
   wb = openpyxl.load_workbook(r"Monroe americana planilla\Monroe.xlsx")
except FileNotFoundError:
   print('El nombre de archivo es incorecto / El archivo indicado no existe')
   print('')
   messagebox.showerror(title='Error', message='El nombre de archivo es incorecto / El archivo indicado no existe')



#me imprime los nombres de las hojas para controlar nada mas
sheetNames = wb.sheetnames
# print(sheetNames) # se podria borrar

#carga la hoja que quiero trabajar
sheet = wb[sheetNames[0]]

#me imprime el valor de la celda D18 para probar nada mas
print(sheet['D18'].value)


# Defino las columnas como variables para luego iterarlas
colA = sheet['A']
colH = sheet['H']
colS = sheet['S']
colP = sheet['P']
colAK = sheet['AK']
colAL = sheet['AL']
colAD = sheet['AD']
colAM = sheet['AM']

for cell in colP[1:]:
   if cell.value != None:
    cellValue = cell.value
    print(cellValue)
    parsedValue = float(cellValue.replace(',',''))
    cell.value = parsedValue

for cell in colS[1:]:
   if cell.value != None:
    cellValue = cell.value
    print(cellValue)
    parsedValue = float(cellValue.replace(',',''))
    cell.value = parsedValue

#cargo los valores que quiero buscar 
per_iva= 'Perc. I.V.A. R.G. 3337 3%'
cell_cabecera = 'Cabecera'
cell_impuesto = 'Impuesto'
per_iibb_3= 'Perc.IB 49/98 Bs As Med. 3%'
per_iibb_25= 'Perc.IB 49/98 BsAs Otros 2,5%'
iva_21 = 'I.V.A. 21%'
iva_105 = 'I.V.A. 10.5%'
iva_27 = 'I.V.A. 27%' #EStimo sera asi por que nunca vino uno como para saber.

sheet['AA1'].value = 'Percep. IVA'
sheet['AB1'].value = 'Percep IIBB 3%'
sheet['AC1'].value = 'Percep IIBB 2.5%'
sheet['AD1'].value = 'Percep IIBB Total'
sheet['AE1'].value = 'IVA 21%'
sheet['AF1'].value = 'Neto 21%'
sheet['AG1'].value = 'IVA 10.5%'
sheet['AH1'].value = 'Neto 10.5%'
sheet['AI1'].value = 'IVA 27%'
sheet['AJ1'].value = 'Neto 27%'
sheet['AK1'].value = 'Sumas Net Grav'
sheet['AL1'].value = 'Exento/No grav'
sheet['AM1'].value = 'CUIT Monroe'


for cell1 in colA:
    if cell1.value == cell_cabecera:
      detailsRow = cell1.row + 1
      for col in sheet.iter_rows(min_row=detailsRow, min_col=14, max_col=14):
         print(col[0]) # agregado para cheque borrar en final
         # counter += 1
         # print('Run:', counter, cell2) # agregado para cheque borrar en final
         if sheet.cell(row=col[0].row, column=1).value == cell_cabecera:
            # asdasd = sheet.cell(row=col[0].row, column=1) #puesto para debuggear, se puede borrar
            # print(asdasd) #puesto para debuggear, se puede borrar
            break
         elif col[0].value == per_iva:
            # print(sheet.cell(row=cell2.row , column=13).value) # agregado para chequeo borrar en final
            cell_iva_copy = sheet.cell(row=col[0].row, column=19).value
            print(cell_iva_copy)
            sheet.cell(row=cell1.row, column=27, value=cell_iva_copy)
            debugVar = sheet.cell(row=cell1.row, column=16).value #voy a guardar esto en una variable para podr debugearlo.
            print(debugVar)
         elif col[0].value == per_iibb_3:
            cell_iib3_copy = sheet.cell(row=col[0].row, column=19).value
            sheet.cell(row=cell1.row, column=28, value=cell_iib3_copy)
         elif col[0].value == per_iibb_25:
            cell_iib25_copy = sheet.cell(row=col[0].row, column=19).value
            sheet.cell(row=cell1.row, column=29, value=cell_iib25_copy)
         elif col[0].value == iva_21:
            cell_iva21_copy = sheet.cell(row=col[0].row, column=19).value
            sheet.cell(row=cell1.row, column=31, value=cell_iva21_copy)
            cell_ngrav21 = cell_iva21_copy/0.21
            sheet.cell(row=cell1.row, column=32, value=cell_ngrav21)
         elif col[0].value == iva_105:
            cell_iva105_copy = sheet.cell(row=col[0].row, column=19).value
            sheet.cell(row=cell1.row, column=33, value=cell_iva105_copy)
            cell_ngrav105 = cell_iva105_copy/0.105
            sheet.cell(row=cell1.row, column=34, value=cell_ngrav105)
         elif col[0].value == iva_27:
            cell_iva27_copy = sheet.cell(row=col[0].row, column=19).value
            sheet.cell(row=cell1.row, column=35, value=cell_iva27_copy)
            cell_ngrav27 = cell_iva27_copy/0.27
            sheet.cell(row=cell1.row, column=36, value=cell_ngrav27)

for cell in colA[1:]:
   if cell.value != cell_cabecera and cell.value is not None:
      sheet.delete_rows(cell.row)




for cell in colAK[1:]:
   if sheet.cell(row=cell.row, column=32).value != None:
      net21 = float(sheet.cell(row=cell.row, column=32).value)
   else: net21 = 0
   if sheet.cell(row=cell.row, column=34).value != None:
      net105 = float(sheet.cell(row=cell.row, column=34).value)
   else: net105 = 0
   if sheet.cell(row=cell.row, column=36).value != None:
      net27 = float(sheet.cell(row=cell.row, column=36).value)
   else: net27 = 0
   cell.value = (net21+net105+net27)

for cell in colAL[1:-1]:
   if sheet.cell(row=cell.row, column=16).value != None:
      sumNetos = sheet.cell(row=cell.row, column=16).value
   else: print('Error al calcular no gravado, columna P tiene valores vacios')
   if sheet.cell(row=cell.row, column=37).value != None:
      netoGrav = sheet.cell(row=cell.row, column=37).value
   else: netoGrav = 0
   cell.value = (sumNetos-netoGrav)

for cell in colAD[1:]:
   if sheet.cell(row=cell.row, column=28).value != None:
      percep3 = sheet.cell(row=cell.row, column=28).value
   else: percep3 = 0
   if sheet.cell(row=cell.row, column=29).value != None:
      percep25 = sheet.cell(row=cell.row, column=29).value
   else: percep25 = 0
   cell.value = (percep3 + percep25)

for cell in colAM[1:]:
   cell.value = '330517059095'



print(sheet.max_row)






#itero por las celdas de la columna H, y si encuentro una celda cuyo valor coincida con la varible que le paso me imprime el valor por consola, lo copia en un variable y dsp graba esa variable en una celda.


#guarda el libro (con otro nombre)
wb.save(r"Monroe americana planilla\MonroeArreglado.xlsx")




## Asi como esta funciona perfecto.

'''
to-do:
-Procesar varias planillas de una (dif quincena o sucursal) [En Standby por que no es tan necesario, parece que hay una planilla con todo junto]
-Mostrar Error cuando no se enuentre la planilla - DONE
-convertir a csv
-Dar la posibilidad de indicar un nuevo nombre para la planilla.
-Automatizar el nombre de salida (con periodo por ej)
-Incorporar controles, ej de totales y etc por si aparece una variable no esperada para no tomar valores erroneos
-empaquetar?

'''