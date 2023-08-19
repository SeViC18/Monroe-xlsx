import openpyxl
import os
import sys
import tkinter
import csv
from tkinter import messagebox
import locale
from locale import atof

#carga la planilla
try:
   wb = openpyxl.load_workbook(r"MonroeArreglado 04-2023.xlsx")
except FileNotFoundError:
   print('El nombre de archivo es incorecto / El archivo indicado no existe')
   print('')
   messagebox.showerror(title='Error', message='El nombre de archivo es incorecto / El archivo indicado no existe')

#Guardo los nombres de las hojas en una variable
sheetNames = wb.sheetnames


#carga la hoja que quiero trabajar
sheet = wb[sheetNames[0]]

loc = locale.getlocale()
print(loc)

with open('Monroe.csv', 'w',newline='') as csvfile:
   locale.setlocale(locale.LC_NUMERIC, 'Spanish_Argentina' )
   csv_writer = csv.writer(csvfile, delimiter=';')
   for row in sheet.iter_rows():
      csvRow = []
      for cell in row:
         value = cell.value
         if isinstance(value, float):
            value = atof(str(value))
         csvRow.append(value)  
      csvfile.write(';'.join(csvRow) + '\n')


#estoy intenando hacer la conversion con locale en base a lo queme tiro chatgpt tengi que seguirlo
#convierte bien pero tiene algun problema con los numeros 