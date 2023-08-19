import openpyxl
import os
import sys
import tkinter
import csv
from tkinter import messagebox
import locale

#carga la planilla
try:
   wb = openpyxl.load_workbook(r"MonroeArreglado 04-2023.xlsx")
except FileNotFoundError:
   print('El nombre de archivo es incorecto / El archivo indicado no existe')
   print('')
   messagebox.showerror(title='Error', message='El nombre de archivo es incorecto / El archivo indicado no existe')

#Guardo los nombres de las hojas en una variable
sheetNames = wb.sheetnames


#carga la hoja que quiero trabajar
sheet = wb[sheetNames[0]]



with open('Monroe.csv', 'w',newline='') as csvfile:
   csv_writer = csv.writer(csvfile, delimiter=';')
   for row in sheet.iter_rows():
      csv_writer.writerow([cell.value for cell in row])



#convierte bien pero tiene algun problema con los numeros 